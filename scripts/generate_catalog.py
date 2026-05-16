from __future__ import annotations

import json
import re
import shutil
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "src"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

from humanoidsim.task_schema import (  # noqa: E402
    Condition,
    EquipmentRequirement,
    EquipmentUseMode,
    ParameterSpec,
    RiskLevel,
    SafetyConstraint,
    StepCall,
    SuccessCriterion,
    TaskLevel,
    TaskSpec,
    ToolRequirement,
    ToolUseMode,
    VehicleRequirement,
    VehicleUseMode,
    dump_json_file,
)

WORKBOOK = Path(r"C:\Github\manufacturing_humanoid_core_task_catalog.xlsx")
if not WORKBOOK.exists():
    WORKBOOK = ROOT / "manufacturing_humanoid_core_task_catalog.xlsx"
ASSET_SOURCE = Path(r"C:\Github\ManSim\replay_studio\public\assets\worker_processed")
DATA = ROOT / "data"
TASKS = DATA / "tasks"
PRIMITIVES = DATA / "primitives"
ASSETS = ROOT / "assets" / "worker_processed"
EXAMPLES = ROOT / "examples"

SEPARATOR = "Ąć"
CATALOG_VERSION = "0.2.0-core"

ACTION_BY_TEMPLATE = {
    "PT-REPLENISH": "TRANSFER",
    "PT-MACHINE": "EXECUTE_MACHINE_ACTION",
    "PT-ASSEMBLY": "EXECUTE_ASSEMBLY_ACTION",
    "PT-MATERIAL": "EXECUTE_MATERIAL_ACTION",
    "PT-PROCESS": "EXECUTE_PROCESS_ACTION",
    "PT-QUALITY": "EXECUTE_QUALITY_ACTION",
    "PT-MAINTENANCE": "EXECUTE_MAINTENANCE_ACTION",
    "PT-EHS": "EXECUTE_EHS_ACTION",
    "PT-PACKAGING": "EXECUTE_PACKAGING_ACTION",
    "PT-WAREHOUSE": "EXECUTE_WAREHOUSE_ACTION",
    "PT-HUMAN": "EXECUTE_HUMAN_COLLABORATION_ACTION",
}

CAPABILITIES_BY_TEMPLATE = {
    "PT-SYSTEM": ["system_operation", "digital_context"],
    "PT-TRANSFER": ["navigation", "object_localization", "manipulation", "payload_handling"],
    "PT-REPLENISH": ["navigation", "manipulation", "inventory_interaction"],
    "PT-VEHICLE": ["navigation", "vehicle_operation", "load_handling"],
    "PT-MACHINE": ["machine_interface", "safety_zone_check", "manipulation"],
    "PT-ASSEMBLY": ["assembly_manipulation", "object_localization", "fine_manipulation"],
    "PT-MATERIAL": ["material_application", "surface_preparation", "tool_use"],
    "PT-PROCESS": ["tool_use", "process_rework", "inspection"],
    "PT-QUALITY": ["inspection", "measurement", "digital_recording"],
    "PT-MAINTENANCE": ["maintenance", "diagnostics", "tool_use", "safety_zone_check"],
    "PT-EHS": ["cleaning", "ehs_audit", "hazard_reporting"],
    "PT-PACKAGING": ["packaging", "labeling", "manipulation"],
    "PT-WAREHOUSE": ["warehouse_operation", "inventory_interaction", "digital_recording"],
    "PT-DIGITAL": ["digital_transaction", "traceability"],
    "PT-HUMAN": ["human_collaboration", "handover", "safe_interaction"],
}

COMPOSITE_STEP_OVERRIDES = {
    "RECOVER_FROM_FAULT": ["CHECK_CONTEXT", "SELF_CHECK", "EXECUTE_SYSTEM_ACTION", "VERIFY_ROBOT_STATE", "LOG_RESULT"],
    "REPLENISH_MATERIAL": ["CHECK_REQUEST", "PRIMITIVE_IDENTIFY_ITEM", "TRANSFER", "VERIFY_LEVEL_OR_QUANTITY", "UPDATE_RECORD"],
    "REMOVE_MATERIAL": ["CHECK_REQUEST", "PRIMITIVE_IDENTIFY_ITEM", "TRANSFER", "VERIFY_LEVEL_OR_QUANTITY", "UPDATE_RECORD"],
    "OPERATE_VEHICLE_TRANSPORT": ["VERIFY_AUTHORIZATION", "TRANSFER", "PARK_OR_RELEASE_VEHICLE", "VERIFY_PLACEMENT"],
    "SETUP_MACHINE": ["CHECK_SAFETY_ZONE", "READ_MACHINE_STATE", "LOAD_MACHINE", "VERIFY_MACHINE_STATE", "LOG_RESULT"],
    "CHANGE_MACHINE_CONFIGURATION": ["CHECK_SAFETY_ZONE", "OPERATE_MACHINE_INTERFACE", "VERIFY_MACHINE_STATE", "LOG_RESULT"],
    "CLEAR_MACHINE_FAULT": ["CHECK_SAFETY_ZONE", "OPERATE_MACHINE_INTERFACE", "VERIFY_MACHINE_STATE", "LOG_RESULT"],
    "ASSEMBLE_COMPONENTS": ["INSERT_COMPONENT", "FASTEN_COMPONENT", "VERIFY_ASSEMBLY"],
    "DISASSEMBLE_COMPONENTS": ["UNFASTEN_COMPONENT", "REMOVE_COMPONENT", "VERIFY_ASSEMBLY"],
    "ROUTE_FLEXIBLE_COMPONENT": ["INSERT_COMPONENT", "CONNECT_COMPONENT", "VERIFY_ASSEMBLY"],
    "CURE_MATERIAL": ["PREPARE_SURFACE", "APPLY_MATERIAL", "VERIFY_MATERIAL_APPLICATION", "RECORD_RESULT"],
    "RUN_TEST": ["IDENTIFY_ITEM", "INSPECT_PRODUCT", "RECORD_QUALITY_RESULT"],
    "SORT_OR_QUARANTINE_ITEM": ["IDENTIFY_ITEM", "INSPECT_PRODUCT", "RECORD_QUALITY_RESULT"],
    "PREVENTIVE_MAINTENANCE": ["CHECK_SAFETY_ZONE", "INSPECT_MACHINE", "LOG_RESULT"],
    "DIAGNOSE_MACHINE": ["CHECK_SAFETY_ZONE", "INSPECT_MACHINE", "LOG_RESULT"],
    "REPAIR_MACHINE": ["CHECK_SAFETY_ZONE", "INSPECT_MACHINE", "EXECUTE_MAINTENANCE_ACTION", "VERIFY_MACHINE_STATE", "LOG_RESULT"],
    "REPLACE_MACHINE_PART": ["CHECK_SAFETY_ZONE", "INSPECT_MACHINE", "EXECUTE_MAINTENANCE_ACTION", "VERIFY_MACHINE_STATE", "LOG_RESULT"],
    "SERVICE_FLUID_OR_LUBRICATION": ["INSPECT_MACHINE", "EXECUTE_MAINTENANCE_ACTION", "VERIFY_MACHINE_STATE", "LOG_RESULT"],
    "CALIBRATE_MACHINE": ["INSPECT_MACHINE", "EXECUTE_MAINTENANCE_ACTION", "VERIFY_MACHINE_STATE", "LOG_RESULT"],
    "CLEAN_AREA": ["CLEAN_ASSET", "VERIFY_AREA_STATE", "REPORT_RESULT"],
    "COLLECT_WASTE_OR_SCRAP": ["TRANSFER", "UPDATE_INVENTORY_RECORD"],
    "RESPOND_TO_SPILL": ["CLEAN_AREA", "REPORT_HAZARD"],
    "CONDUCT_EHS_OR_5S_AUDIT": ["CLEAN_ASSET", "REPORT_HAZARD"],
    "PACK_PRODUCT": ["LABEL_ITEM_OR_PACKAGE", "VERIFY_PACKAGE"],
    "UNPACK_MATERIAL": ["VERIFY_PACKAGE", "UPDATE_INVENTORY_RECORD"],
    "UNITIZE_LOAD": ["PACK_PRODUCT", "VERIFY_PACKAGE"],
    "RECEIVE_MATERIAL": ["IDENTIFY_ITEM", "UPDATE_INVENTORY_RECORD"],
    "PUTAWAY_ITEM": ["TRANSFER", "UPDATE_INVENTORY_RECORD"],
    "PICK_INVENTORY": ["COUNT_INVENTORY", "TRANSFER", "UPDATE_INVENTORY_RECORD"],
    "BUILD_KIT": ["PICK_INVENTORY", "UPDATE_INVENTORY_RECORD"],
    "FETCH_FOR_OPERATOR": ["TRANSFER", "HANDOVER_ITEM"],
    "ASSIST_OPERATOR_MOVE_OR_LIFT": ["TRANSFER", "HANDOVER_ITEM"],
}

ATOMIC_STEP_OVERRIDES = {
    "INSPECT_MACHINE": ["CHECK_SAFETY_ZONE", "VERIFY_LOCKOUT_IF_REQUIRED", "INSPECT_OR_DIAGNOSE", "LOG_RESULT"],
}

SPECIAL_CHILD_ARG_MAP = {
    ("REPLENISH_MATERIAL", "TRANSFER"): {"item": "$inputs.item", "source": "$inputs.source", "destination": "$inputs.destination"},
    ("REMOVE_MATERIAL", "TRANSFER"): {"item": "$inputs.item", "source": "$inputs.source", "destination": "$inputs.destination"},
    ("SETUP_MACHINE", "LOAD_MACHINE"): {
        "machine": "$inputs.machine",
        "item": {"derived_from_parent": "SETUP_MACHINE", "field": "item"},
        "source": {"derived_from_parent": "SETUP_MACHINE", "field": "source"},
        "target_slot": {"derived_from_parent": "SETUP_MACHINE", "field": "target_slot"},
    },
    ("REPAIR_MACHINE", "INSPECT_MACHINE"): {"machine": "$inputs.machine", "inspection_plan": "$inputs.fault"},
    ("DIAGNOSE_MACHINE", "INSPECT_MACHINE"): {"machine": "$inputs.machine", "inspection_plan": "$inputs.symptom"},
    ("PREVENTIVE_MAINTENANCE", "INSPECT_MACHINE"): {"machine": "$inputs.asset", "inspection_plan": "$inputs.checklist"},
    ("FETCH_FOR_OPERATOR", "TRANSFER"): {
        "item": "$inputs.request",
        "source": {"derived_from_parent": "FETCH_FOR_OPERATOR", "field": "source"},
        "destination": "$inputs.operator_or_station",
    },
    ("FETCH_FOR_OPERATOR", "HANDOVER_ITEM"): {
        "item": "$inputs.request",
        "recipient": "$inputs.operator_or_station",
        "handover_spec": "$inputs.request",
    },
    ("ASSIST_OPERATOR_MOVE_OR_LIFT", "TRANSFER"): {"item": "$inputs.item", "source": "$inputs.source", "destination": "$inputs.destination"},
    ("ASSIST_OPERATOR_MOVE_OR_LIFT", "HANDOVER_ITEM"): {
        "item": "$inputs.item",
        "recipient": "$inputs.operator",
        "handover_spec": "$inputs.assist_spec",
    },
    ("PUTAWAY_ITEM", "TRANSFER"): {
        "item": "$inputs.item",
        "source": {"derived_from_parent": "PUTAWAY_ITEM", "field": "source"},
        "destination": "$inputs.storage_location",
    },
    ("PICK_INVENTORY", "TRANSFER"): {
        "item": "$inputs.request",
        "source": {"derived_from_parent": "PICK_INVENTORY", "field": "source"},
        "destination": {"derived_from_parent": "PICK_INVENTORY", "field": "destination"},
    },
    ("COLLECT_WASTE_OR_SCRAP", "TRANSFER"): {
        "item": "$inputs.waste_or_scrap",
        "source": "$inputs.source",
        "destination": "$inputs.disposal_location",
    },
}


def main() -> int:
    for path in (TASKS, PRIMITIVES, ASSETS, EXAMPLES):
        path.mkdir(parents=True, exist_ok=True)
    for folder in (TASKS, PRIMITIVES):
        for path in folder.glob("*.json"):
            path.unlink()
    _copy_assets()

    wb = load_workbook(WORKBOOK, data_only=True)
    templates = _load_templates(wb)
    rows = _load_task_rows(wb)
    task_codes = {str(row["Task Code"]) for row in rows}
    task_levels = {
        str(row["Task Code"]): (
            TaskLevel.ATOMIC_TASK if str(row["Suggested Level"]).upper().startswith("ATOMIC") else TaskLevel.COMPOSITE_TASK
        )
        for row in rows
    }
    task_inputs = {str(row["Task Code"]): _split_csv(str(row.get("Primary Inputs", ""))) for row in rows}
    primitive_codes: dict[str, set[str]] = {}
    task_paths: list[str] = []

    for row in rows:
        spec, used_primitives = _task_from_row(row, templates, task_codes, task_levels, task_inputs)
        for code, args in used_primitives.items():
            primitive_codes.setdefault(code, set()).update(args)
        rel = f"data/tasks/{row['Task No']}_{row['Task Code']}.json"
        dump_json_file(spec, str(ROOT / rel))
        task_paths.append(rel)

    primitive_paths: list[str] = []
    for code in sorted(primitive_codes):
        spec = TaskSpec(
            code=code,
            level=TaskLevel.PRIMITIVE_SKILL,
            version="1.0.0",
            name=_title(code),
            description=f"Primitive skill generated from manufacturing humanoid primitive templates for {code}.",
            inputs=[ParameterSpec(name=name, type_hint="Any", required=False) for name in sorted(primitive_codes[code])],
            outputs=[ParameterSpec(name="result", type_hint="dict", required=False)],
            success_criteria=[SuccessCriterion(metric="status", operator="==", value="SUCCESS")],
            metadata={"generated_from": "primitive_templates"},
        )
        rel = f"data/primitives/{code}.json"
        dump_json_file(spec, str(ROOT / rel))
        primitive_paths.append(rel)

    dump_json_file(_templates_json(templates), str(DATA / "primitive_templates.json"))
    index = {
        "schema_version": "0.1.0",
        "catalog_version": CATALOG_VERSION,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source_workbook": str(WORKBOOK),
        "task_count": len(rows),
        "primitive_count": len(primitive_paths),
        "task_paths": task_paths,
        "primitive_paths": primitive_paths,
        "task_codes": [row["Task Code"] for row in rows],
    }
    dump_json_file(index, str(DATA / "task_catalog_core.json"))
    _write_example_sequence(rows)
    print(json.dumps({"task_count": len(rows), "primitive_count": len(primitive_paths)}, indent=2))
    return 0


def _copy_assets() -> None:
    for path in ASSET_SOURCE.glob("*.png"):
        shutil.copy2(path, ASSETS / path.name)


def _load_templates(wb: Any) -> dict[str, dict[str, Any]]:
    ws = wb["Primitive Templates"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    templates: dict[str, dict[str, Any]] = {}
    for r in range(2, ws.max_row + 1):
        row = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        template_id = row.get("Template ID")
        if not template_id:
            continue
        normalized_steps = [_normalize_primitive(step, str(template_id)) for step in _split_sequence(str(row["Primitive Sequence"]))]
        templates[str(template_id)] = {
            "template_id": str(template_id),
            "name": str(row.get("Template Name", "")),
            "raw_sequence": str(row.get("Primitive Sequence", "")),
            "used_for": str(row.get("Used For", "")),
            "normalized_steps": normalized_steps,
        }
    return templates


def _load_task_rows(wb: Any) -> list[dict[str, Any]]:
    ws = wb["Core Task Catalog"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        row = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        if row.get("Task Code"):
            rows.append(row)
    return rows


def _task_from_row(
    row: dict[str, Any],
    templates: dict[str, dict[str, Any]],
    task_codes: set[str],
    task_levels: dict[str, TaskLevel],
    task_inputs: dict[str, list[str]],
) -> tuple[TaskSpec, dict[str, set[str]]]:
    template_id = str(row["Primitive Template ID"])
    template = templates[template_id]
    parent_inputs = _split_csv(str(row.get("Primary Inputs", "")))
    inputs = [_parameter(name) for name in parent_inputs]
    tools, optional_tools = _tools(row.get("Required Tools"))
    vehicles, optional_vehicles = _vehicles(row.get("Required Vehicles"))
    equipment, optional_equipment = _equipment(row.get("Required Equipment / Assets"))
    level = task_levels[str(row["Task Code"])]
    steps, primitives = _steps(
        str(row["Task Code"]),
        template,
        task_codes,
        task_levels,
        parent_inputs,
        task_inputs,
        level,
    )
    animation = _animation(row)
    capabilities = list(dict.fromkeys(CAPABILITIES_BY_TEMPLATE.get(template_id, []) + _resource_capabilities(tools, vehicles, equipment, row)))

    spec = TaskSpec(
        code=str(row["Task Code"]),
        level=level,
        version=CATALOG_VERSION,
        name=str(row["Task Name"]),
        description=str(row["Purpose"]),
        inputs=inputs,
        outputs=[ParameterSpec(name="result", type_hint="dict", required=False, description="Task execution result payload.")],
        required_capabilities=capabilities,
        required_tools=tools,
        required_vehicles=vehicles,
        required_equipment=equipment,
        safety=_safety(row),
        preconditions=[Condition(expression="required_inputs_present", description="All required TaskInstance args must be provided.")],
        postconditions=[Condition(expression="execution_record_written", description="Execution result is recorded for traceability.")],
        success_criteria=[SuccessCriterion(metric="status", operator="==", value="SUCCESS", description="Task completes successfully.")],
        steps=steps,
        metadata={
            "catalog": {
                "task_no": row["Task No"],
                "category_id": row["Category ID"],
                "category": row["Category"],
                "function_signature": row["Function Signature"],
                "implementation_priority": row["Implementation Priority"],
                "customization_notes": row["Customization / Implementation Notes"],
                "primitive_template_id": template_id,
            },
            "primitive_template": _effective_template_metadata(template, steps),
            "resources": {
                "raw_tools": row.get("Required Tools"),
                "raw_vehicles": row.get("Required Vehicles"),
                "raw_equipment": row.get("Required Equipment / Assets"),
                "optional_aliases": {
                    "tools": optional_tools,
                    "vehicles": optional_vehicles,
                    "equipment": optional_equipment,
                },
            },
            "animation": {"frames": animation, "placeholder": True, "source": "ManSim replay_studio worker_processed"},
        },
    )
    return spec, primitives


def _split_sequence(value: str) -> list[str]:
    text = value.replace("→", SEPARATOR).replace("->", SEPARATOR)
    return [item.strip() for item in text.split(SEPARATOR) if item.strip()]


def _effective_template_metadata(template: dict[str, Any], steps: list[StepCall]) -> dict[str, Any]:
    metadata = dict(template)
    metadata["normalized_steps"] = [
        {
            "raw": step.call_code,
            "call_code": step.call_code,
            "expected_level": step.expected_level.value if step.expected_level else None,
            "args": step.args,
        }
        for step in steps
    ]
    return metadata


def _normalize_primitive(raw: str, template_id: str) -> dict[str, Any]:
    call_text = raw.strip()
    if " / " in call_text or re.search(r"\bor\b", call_text, flags=re.IGNORECASE):
        code = ACTION_BY_TEMPLATE.get(template_id, _code(call_text))
        args: dict[str, str] = {}
    else:
        match = re.match(r"^([A-Z0-9_]+)\((.*)\)$", call_text.strip())
        if match:
            code = match.group(1)
            args = {name.strip(): f"$inputs.{name.strip()}" for name in match.group(2).split(",") if name.strip()}
        else:
            code = _code(call_text)
            args = {}
    return {"raw": raw, "call_code": code, "args": args}


def _seed_template_primitives(templates: dict[str, dict[str, Any]], task_codes: set[str]) -> dict[str, set[str]]:
    primitives: dict[str, set[str]] = {}
    for template in templates.values():
        for row in template["normalized_steps"]:
            raw_code = row["call_code"]
            code = f"PRIMITIVE_{raw_code}" if raw_code in task_codes else raw_code
            primitives.setdefault(code, set()).update(row.get("args", {}).keys())
    return primitives


def _steps(
    task_code: str,
    template: dict[str, Any],
    task_codes: set[str],
    task_levels: dict[str, TaskLevel],
    parent_inputs: list[str],
    task_inputs: dict[str, list[str]],
    task_level: TaskLevel,
) -> tuple[list[StepCall], dict[str, set[str]]]:
    steps: list[StepCall] = []
    primitives: dict[str, set[str]] = {}
    seen: dict[str, int] = {}
    previous: str | None = None
    if task_code in ATOMIC_STEP_OVERRIDES:
        source_rows = [{"call_code": code, "args": {}} for code in ATOMIC_STEP_OVERRIDES[task_code]]
    elif task_level == TaskLevel.COMPOSITE_TASK and task_code in COMPOSITE_STEP_OVERRIDES:
        source_rows = [{"call_code": code, "args": {}} for code in COMPOSITE_STEP_OVERRIDES[task_code]]
    else:
        source_rows = template["normalized_steps"]

    for index, row in enumerate(source_rows, start=1):
        raw_code = row["call_code"]
        if task_level == TaskLevel.ATOMIC_TASK and raw_code in task_codes:
            code = f"PRIMITIVE_{raw_code}"
        else:
            code = raw_code

        expected_level = task_levels.get(code, TaskLevel.PRIMITIVE_SKILL)
        args = _step_args(task_code, code, row.get("args", {}), expected_level, parent_inputs, task_inputs)
        if expected_level == TaskLevel.PRIMITIVE_SKILL:
            primitives.setdefault(code, set()).update(args.keys())
        seen[code] = seen.get(code, 0) + 1
        suffix = f"_{seen[code]}" if seen[code] > 1 else ""
        step_id = f"s{index:02d}_{code.lower()}{suffix}"
        steps.append(
            StepCall(
                step_id=step_id,
                call_code=code,
                args=args,
                expected_level=expected_level,
                depends_on=[previous] if previous else [],
            )
        )
        previous = step_id
    return steps, primitives


def _step_args(
    parent_code: str,
    call_code: str,
    template_args: dict[str, Any],
    expected_level: TaskLevel,
    parent_inputs: list[str],
    task_inputs: dict[str, list[str]],
) -> dict[str, Any]:
    if expected_level == TaskLevel.PRIMITIVE_SKILL:
        return dict(template_args)

    if (parent_code, call_code) in SPECIAL_CHILD_ARG_MAP:
        return dict(SPECIAL_CHILD_ARG_MAP[(parent_code, call_code)])

    output: dict[str, Any] = {}
    for name in task_inputs.get(call_code, []):
        if name in parent_inputs:
            output[name] = f"$inputs.{name}"
        else:
            output[name] = {"derived_from_parent": parent_code, "field": name}
    return output


def _parameter(name: str) -> ParameterSpec:
    return ParameterSpec(name=name, type_hint=_type_hint(name), required=True, description=f"Catalog input parameter: {name}.")


def _type_hint(name: str) -> str:
    lowered = name.lower()
    if any(token in lowered for token in ("source", "destination", "station", "area", "location", "dock", "path")):
        return "LocationRef | str"
    if any(token in lowered for token in ("item", "part", "product", "machine", "robot", "tool", "asset", "entity")):
        return "EntityRef | str"
    if "target_soc" in lowered:
        return "float"
    return "Any"


def _tools(value: Any) -> tuple[list[ToolRequirement], list[str]]:
    aliases, optional = _resource_aliases(value)
    return [
        ToolRequirement(
            alias=alias,
            tool_type=resource_type,
            use_mode=_tool_mode(resource_type),
            settings={"source_text": str(value or "")},
            safety_notes="Optional resource from catalog." if alias in optional else "",
        )
        for alias, resource_type in aliases
    ], sorted(optional)


def _vehicles(value: Any) -> tuple[list[VehicleRequirement], list[str]]:
    aliases, optional = _resource_aliases(value)
    return [
        VehicleRequirement(
            alias=alias,
            vehicle_type=resource_type,
            use_mode=VehicleUseMode.LIFT_AND_TRANSPORT if any(token in resource_type for token in ("forklift", "pallet_jack")) else VehicleUseMode.OPERATE,
            safety_notes="Optional resource from catalog." if alias in optional else "",
        )
        for alias, resource_type in aliases
    ], sorted(optional)


def _equipment(value: Any) -> tuple[list[EquipmentRequirement], list[str]]:
    aliases, optional = _resource_aliases(value)
    return [
        EquipmentRequirement(
            alias=alias,
            equipment_type=resource_type,
            use_mode=_equipment_mode(resource_type),
            settings={"source_text": str(value or "")},
            safety_notes="Optional resource from catalog." if alias in optional else "",
        )
        for alias, resource_type in aliases
    ], sorted(optional)


def _resource_aliases(value: Any) -> tuple[list[tuple[str, str]], set[str]]:
    if value is None or str(value).strip() in {"", "-"}:
        return [], set()
    raw = str(value)
    optional_all = "optional" in raw.lower()
    cleaned = re.sub(r"\boptional\b", "", raw, flags=re.IGNORECASE)
    parts = _split_csv_slash(cleaned)
    pairs: list[tuple[str, str]] = []
    optional: set[str] = set()
    used: set[str] = set()
    for part in parts:
        resource_type = _slug(part)
        if not resource_type or resource_type in used:
            continue
        used.add(resource_type)
        alias = resource_type
        pairs.append((alias, resource_type))
        if optional_all:
            optional.add(alias)
    return pairs, optional


def _split_csv(value: str) -> list[str]:
    return [item.strip() for item in re.split(r",", value) if item and item.strip()]


def _split_csv_slash(value: str) -> list[str]:
    return [item.strip() for item in re.split(r"[,/]", value) if item and item.strip()]


def _tool_mode(resource_type: str) -> ToolUseMode:
    if any(token in resource_type for token in ("scanner", "rfid", "camera")):
        return ToolUseMode.SCANNING
    if any(token in resource_type for token in ("torque", "nutrunner", "rivet", "screw")):
        return ToolUseMode.FASTENING
    if any(token in resource_type for token in ("cutter", "knife", "drill", "tap", "reamer")):
        return ToolUseMode.CUTTING
    if any(token in resource_type for token in ("gauge", "scale", "sensor")):
        return ToolUseMode.MEASURING
    if any(token in resource_type for token in ("wiper", "vacuum", "mop", "brush")):
        return ToolUseMode.CLEANING
    return ToolUseMode.GENERAL


def _equipment_mode(resource_type: str) -> EquipmentUseMode:
    if any(token in resource_type for token in ("machine", "hmi", "controller", "panel")):
        return EquipmentUseMode.MACHINE_INTERFACE
    if any(token in resource_type for token in ("fixture", "jig")):
        return EquipmentUseMode.FIXTURE
    if any(token in resource_type for token in ("tester", "test")):
        return EquipmentUseMode.TESTER
    if any(token in resource_type for token in ("storage", "wms", "erp", "mes")):
        return EquipmentUseMode.STORAGE
    return EquipmentUseMode.OTHER


def _resource_capabilities(
    tools: list[ToolRequirement],
    vehicles: list[VehicleRequirement],
    equipment: list[EquipmentRequirement],
    row: dict[str, Any],
) -> list[str]:
    caps: list[str] = []
    if tools:
        caps.append("tool_use")
    if vehicles:
        caps.append("vehicle_operation")
    if equipment:
        caps.append("equipment_interaction")
    if str(row.get("Safety Risk", "")).upper() == "HIGH":
        caps.append("high_risk_task")
    return caps


def _safety(row: dict[str, Any]) -> SafetyConstraint:
    risk_text = str(row.get("Safety Risk", "LOW")).upper()
    risk = RiskLevel.HIGH if risk_text == "HIGH" else RiskLevel.MEDIUM if risk_text == "MEDIUM" else RiskLevel.LOW
    return SafetyConstraint(
        risk_level=risk,
        human_clearance_required=str(row.get("Category ID")) == "M",
        lockout_tagout_required=str(row.get("Category ID")) == "H" and risk == RiskLevel.HIGH,
        ppe_required=["standard_factory_ppe"] if risk != RiskLevel.LOW else [],
        notes=f"Catalog safety risk: {risk.value}.",
    )


def _animation(row: dict[str, Any]) -> list[str]:
    code = str(row["Task Code"])
    category_id = str(row["Category ID"])
    if code == "OPERATE_VEHICLE_TRANSPORT":
        frames = ("Walk1.png", "Walk2.png")
    elif code in {"LOAD_MACHINE", "UNLOAD_MACHINE", "LOAD_UNLOAD_TRANSFER_INTERFACE"}:
        frames = ("unload1.png", "unload2.png")
    elif category_id in {"B", "K"} or code in {"FETCH_FOR_OPERATOR", "ASSIST_OPERATOR_MOVE_OR_LIFT"}:
        frames = ("Delivery1.png", "Delivery2.png")
    elif category_id in {"H", "I"}:
        frames = ("Fix1.png", "Fix2.png")
    else:
        frames = ("setup1.png", "setup2.png")
    return [f"assets/worker_processed/{name}" for name in frames]


def _templates_json(templates: dict[str, dict[str, Any]]) -> dict[str, Any]:
    return {"catalog_version": CATALOG_VERSION, "templates": templates}


def _write_example_sequence(rows: list[dict[str, Any]]) -> None:
    selected = [
        "INITIALIZE_ROBOT",
        "SELF_CHECK",
        "LOAD_WORK_CONTEXT",
        "TRANSFER",
        "LOAD_MACHINE",
        "INSPECT_PRODUCT",
        "SORT_OR_QUARANTINE_ITEM",
        "COMPLETE_WORK_ORDER",
    ]
    by_code = {row["Task Code"]: row for row in rows}
    all_caps = sorted({cap for row in rows for cap in CAPABILITIES_BY_TEMPLATE.get(str(row["Primitive Template ID"]), [])} | {"*", "tool_use", "vehicle_operation", "equipment_interaction", "high_risk_task"})
    tasks = []
    for index, code in enumerate(selected, start=1):
        row = by_code[code]
        tasks.append(
            {
                "instance_id": f"TASK-{index:04d}",
                "task_code": code,
                "work_order_id": "WO-DEMO-0001",
                "assigned_robot_id": "HUMANOID-01",
                "priority": 10 - index,
                "args": {name: _sample_arg(name) for name in _split_csv(str(row.get("Primary Inputs", "")))},
            }
        )
    payload = {
        "humanoids": [
            {
                "humanoid_id": "HUMANOID-01",
                "capabilities": all_caps,
                "max_payload_kg": 35.0,
                "supported_tools": ["*"],
                "supported_vehicles": ["*"],
                "supported_equipment": ["*"],
            }
        ],
        "tasks": tasks,
    }
    (EXAMPLES / "manufacturing_sequence.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _sample_arg(name: str) -> Any:
    lowered = name.lower()
    if any(token in lowered for token in ("source", "destination", "station", "area", "location")):
        return {"location_id": f"demo/{lowered}"}
    if any(token in lowered for token in ("item", "part", "product", "machine", "robot", "tool", "entity")):
        return {"entity_type": lowered, "entity_id": f"DEMO-{_slug(name).upper()}", "weight_kg": 5.0}
    if "target_soc" in lowered:
        return 0.8
    return f"demo_{_slug(name)}"


def _code(value: str) -> str:
    return _slug(value).upper()


def _slug(value: str) -> str:
    text = re.sub(r"[^A-Za-z0-9]+", "_", str(value).strip()).strip("_")
    return re.sub(r"_+", "_", text).lower()


def _title(code: str) -> str:
    return code.replace("_", " ").title()


if __name__ == "__main__":
    raise SystemExit(main())
